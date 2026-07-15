from dataclasses import dataclass
from pathlib import Path

import openpyxl

from app.domain.errors import ManualReviewRequired, ModelRowNotFound
from app.domain.models import InventoryRecord
from app.ingest.model_number import normalize_model


@dataclass(frozen=True)
class InventoryValuePresence:
    price_present: bool
    stock_present: bool


def _is_present(value: object) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def inspect_inventory_values(workbook_path: Path, model: str) -> InventoryValuePresence:
    wanted = normalize_model(model)
    book = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = book.active
        for raw_model, price, stock, *_ in sheet.iter_rows(min_row=2, values_only=True):
            if raw_model is None or normalize_model(str(raw_model)) != wanted:
                continue
            return InventoryValuePresence(
                price_present=_is_present(price),
                stock_present=_is_present(stock),
            )
    finally:
        book.close()
    raise ModelRowNotFound(f"model row does not exist: {wanted}")


def load_inventory(workbook_path: Path, model: str) -> InventoryRecord:
    wanted = normalize_model(model)
    book = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = book.active
        for raw_model, price, stock, *_ in sheet.iter_rows(min_row=2, values_only=True):
            if raw_model is None or normalize_model(str(raw_model)) != wanted:
                continue
            missing = [
                name
                for name, value in (("价格", price), ("库存", stock))
                if not _is_present(value)
            ]
            if missing:
                raise ManualReviewRequired(
                    f"{wanted} 的{'和'.join(missing)}不能为空；请填写真实值后重试"
                )
            try:
                return InventoryRecord(model=wanted, price=int(price), stock=int(stock))
            except (TypeError, ValueError) as error:
                raise ManualReviewRequired(
                    f"{wanted} 的价格和库存必须是有效数字；未修改原表"
                ) from error
    finally:
        book.close()
    raise ModelRowNotFound(f"model row does not exist: {wanted}")
