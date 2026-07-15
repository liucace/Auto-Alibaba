from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from app.domain.errors import ManualReviewRequired
from app.ingest.model_number import exact_model_match, model_folder_key, normalize_model

HEADERS = ("型号", "价格", "库存")
SOURCE_IMAGE_SUFFIXES = frozenset(
    {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}
)


@dataclass(frozen=True)
class InputRequirement:
    key: str
    path: str
    purpose: str
    action: str
    ready: bool


@dataclass(frozen=True)
class ProductInputResult:
    ok: bool
    status: str
    model: str
    folder_key: str
    created: tuple[str, ...]
    checks: dict[str, object]
    requirements: tuple[InputRequirement, ...]
    message: str

    def to_dict(self) -> dict[str, object]:
        return {
            "ok": self.ok,
            "status": self.status,
            "model": self.model,
            "folder_key": self.folder_key,
            "created": list(self.created),
            "checks": self.checks,
            "requirements": [asdict(item) for item in self.requirements],
            "message": self.message,
        }


def _save_workbook(book: openpyxl.Workbook, path: Path) -> None:
    temporary = path.with_name(f"{path.stem}.tmp{path.suffix}")
    try:
        book.save(temporary)
        temporary.replace(path)
    except OSError as error:
        temporary.unlink(missing_ok=True)
        raise ManualReviewRequired(f"库存表无法保存，未覆盖原文件: {path}") from error


def _ensure_inventory_row(workbook: Path, model: str) -> bool:
    if not workbook.exists():
        book = openpyxl.Workbook()
        try:
            sheet = book.active
            sheet.title = "库存"
            sheet.append(list(HEADERS))
            sheet.append([model, None, None])
            _save_workbook(book, workbook)
        finally:
            book.close()
        return True

    try:
        book = openpyxl.load_workbook(workbook)
    except (OSError, ValueError, BadZipFile, InvalidFileException) as error:
        raise ManualReviewRequired(f"库存表无法读取，未覆盖原文件: {workbook}") from error

    try:
        sheet = book.active
        headers = tuple(sheet.cell(row=1, column=index).value for index in range(1, 4))
        if headers != HEADERS:
            raise ManualReviewRequired(
                f"库存表表头必须为 型号、价格、库存，未修改原文件: {workbook}"
            )
        has_model = any(
            raw_model is not None and exact_model_match(str(raw_model), model)
            for raw_model, *_ in sheet.iter_rows(min_row=2, values_only=True)
        )
        if has_model:
            return False
        sheet.append([model, None, None])
        _save_workbook(book, workbook)
        return True
    finally:
        book.close()


def initialize_product_inputs(root: Path, model: str) -> ProductInputResult:
    root = root.resolve()
    normalized = normalize_model(model)
    folder_key = model_folder_key(normalized)
    workbook = root / "price_inventory.xlsx"
    source = root / "data" / "draft_saved" / folder_key
    created: list[str] = []

    workbook.parent.mkdir(parents=True, exist_ok=True)
    inventory_needs_review = _ensure_inventory_row(workbook, normalized)
    if inventory_needs_review:
        created.append(str(workbook.resolve()))

    if not source.is_dir():
        source.mkdir(parents=True, exist_ok=True)
        created.append(str(source.resolve()))

    pdf_count = sum(
        1
        for path in source.iterdir()
        if path.is_file() and path.suffix.lower() == ".pdf" and path.stat().st_size > 0
    )
    image_count = sum(
        1
        for path in source.iterdir()
        if path.is_file()
        and path.suffix.lower() in SOURCE_IMAGE_SUFFIXES
        and path.stat().st_size > 0
    )
    source_ready = pdf_count >= 1 and image_count >= 4
    ready = not created and not inventory_needs_review and source_ready
    requirements = (
        InputRequirement(
            key="inventory",
            path=str(workbook),
            purpose="提供当前完整型号的1688价格和库存。",
            action="打开表格，在当前完整型号行填写真实价格和真实库存；两项都不能留空。",
            ready=not inventory_needs_review,
        ),
        InputRequirement(
            key="source_files",
            path=str(source),
            purpose="保存当前型号的原始PDF规格书和真实产品照片。",
            action="放入至少一份包含完整型号的PDF，以及至少四张当前型号产品照片。",
            ready=source_ready,
        ),
    )
    return ProductInputResult(
        ok=ready,
        status="READY" if ready else "NEEDS_INPUT",
        model=normalized,
        folder_key=folder_key,
        created=tuple(created),
        checks={
            "inventory_workbook": True,
            "inventory_model": True,
            "source_directory": True,
            "pdf_files": pdf_count,
            "source_images": image_count,
        },
        requirements=requirements,
        message=(
            "商品资料已齐全，可以继续准备和上传。"
            if ready
            else "已创建或发现待补充资料；请按requirements核对后重新运行。"
        ),
    )
