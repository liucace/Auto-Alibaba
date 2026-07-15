from pathlib import Path

import openpyxl
import pytest
from PIL import Image

from app.domain.errors import ManualReviewRequired
from app.products.input_onboarding import initialize_product_inputs

MODEL = "W3G800-KS39-03/F01"


def _write_workbook(path: Path, rows: list[tuple[object, object, object]]) -> None:
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["型号", "价格", "库存"])
    for row in rows:
        sheet.append(list(row))
    book.save(path)
    book.close()


def test_missing_inputs_create_inventory_template_and_model_folder(tmp_path: Path) -> None:
    result = initialize_product_inputs(tmp_path, MODEL)

    workbook = tmp_path / "price_inventory.xlsx"
    source = tmp_path / "data" / "draft_saved" / "W3G800-KS39-03F01"
    assert workbook.is_file()
    assert source.is_dir()
    assert result.status == "NEEDS_INPUT"
    assert result.ok is False
    assert str(workbook.resolve()) in result.created
    assert str(source.resolve()) in result.created

    book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    try:
        rows = list(book.active.iter_rows(values_only=True))
    finally:
        book.close()
    assert rows == [("型号", "价格", "库存"), (MODEL, None, None)]

    requirements = {item.key: item for item in result.requirements}
    assert "1688价格和库存" in requirements["inventory"].purpose
    assert "真实价格" in requirements["inventory"].action
    assert "不能留空" in requirements["inventory"].action
    assert "10000" not in requirements["inventory"].action
    assert "PDF" in requirements["source_files"].action
    assert "四张" in requirements["source_files"].action


def test_existing_workbook_appends_model_once_and_preserves_rows(tmp_path: Path) -> None:
    workbook = tmp_path / "price_inventory.xlsx"
    _write_workbook(workbook, [("EXISTING-01", 88, 9)])

    first = initialize_product_inputs(tmp_path, MODEL)
    second = initialize_product_inputs(tmp_path, MODEL)

    book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    try:
        rows = list(book.active.iter_rows(values_only=True))
    finally:
        book.close()
    assert rows == [
        ("型号", "价格", "库存"),
        ("EXISTING-01", 88, 9),
        (MODEL, None, None),
    ]
    assert str(workbook.resolve()) in first.created
    assert first.status == "NEEDS_INPUT"
    assert second.created == ()


def test_ready_requires_top_level_pdf_and_four_source_images(tmp_path: Path) -> None:
    workbook = tmp_path / "price_inventory.xlsx"
    _write_workbook(workbook, [(MODEL, 100, 6)])
    source = tmp_path / "data" / "draft_saved" / "W3G800-KS39-03F01"
    source.mkdir(parents=True)
    (source / "spec.pdf").write_bytes(b"%PDF-input")
    for index in range(4):
        Image.new("RGB", (10, 10)).save(source / f"photo-{index}.jpg")
    optimized = source / "upload_optimized"
    optimized.mkdir()
    Image.new("RGB", (10, 10)).save(optimized / "derived.jpg")

    result = initialize_product_inputs(tmp_path, MODEL)

    assert result.status == "READY"
    assert result.ok is True
    assert result.checks["pdf_files"] == 1
    assert result.checks["source_images"] == 4


def test_derived_images_do_not_satisfy_source_photo_requirement(tmp_path: Path) -> None:
    initialize_product_inputs(tmp_path, MODEL)
    source = tmp_path / "data" / "draft_saved" / "W3G800-KS39-03F01"
    (source / "spec.pdf").write_bytes(b"%PDF-input")
    optimized = source / "upload_optimized"
    optimized.mkdir()
    for index in range(4):
        Image.new("RGB", (10, 10)).save(optimized / f"derived-{index}.jpg")

    result = initialize_product_inputs(tmp_path, MODEL)

    assert result.status == "NEEDS_INPUT"
    assert result.checks["source_images"] == 0


def test_invalid_inventory_workbook_is_not_overwritten(tmp_path: Path) -> None:
    workbook = tmp_path / "price_inventory.xlsx"
    original = b"not-an-xlsx"
    workbook.write_bytes(original)

    with pytest.raises(ManualReviewRequired, match="无法读取"):
        initialize_product_inputs(tmp_path, MODEL)

    assert workbook.read_bytes() == original


def test_incompatible_inventory_headers_are_not_overwritten(tmp_path: Path) -> None:
    workbook = tmp_path / "price_inventory.xlsx"
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["产品", "单价", "数量"])
    sheet.append(["EXISTING-01", 88, 9])
    book.save(workbook)
    book.close()
    original = workbook.read_bytes()

    with pytest.raises(ManualReviewRequired, match="表头"):
        initialize_product_inputs(tmp_path, MODEL)

    assert workbook.read_bytes() == original
