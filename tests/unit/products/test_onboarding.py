from pathlib import Path

import openpyxl

from app.products.onboarding import onboard_product

MODEL = "REAL-AC-FAN/01"
FOLDER_KEY = "REAL-AC-FAN01"


def _write_workbook(path: Path, price: object, stock: object) -> None:
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["型号", "价格", "库存"])
    sheet.append([MODEL, price, stock])
    book.save(path)
    book.close()


def _add_source_files(root: Path, *, pdfs: int = 1, images: int = 4) -> Path:
    source = root / "data" / "draft_saved" / FOLDER_KEY
    source.mkdir(parents=True, exist_ok=True)
    for index in range(pdfs):
        (source / f"spec-{index}.pdf").write_bytes(b"%PDF-real-input")
    for index in range(images):
        (source / f"photo-{index}.jpg").write_bytes(b"real-image-input")
    return source


def test_no_model_is_read_only_and_requests_only_model(tmp_path: Path) -> None:
    result = onboard_product(tmp_path, None)

    assert result.status == "NEEDS_MODEL"
    assert result.model is None
    assert result.created == ()
    assert "完整型号" in result.next_action
    assert list(tmp_path.iterdir()) == []


def test_blank_model_is_also_read_only(tmp_path: Path) -> None:
    result = onboard_product(tmp_path, "   ")

    assert result.status == "NEEDS_MODEL"
    assert list(tmp_path.iterdir()) == []


def test_real_model_creates_only_exact_row_and_folder(tmp_path: Path) -> None:
    result = onboard_product(tmp_path, MODEL)

    assert result.status == "NEEDS_PRICE_STOCK"
    assert result.model == MODEL
    assert result.folder_key == FOLDER_KEY
    assert (tmp_path / "price_inventory.xlsx").is_file()
    assert (tmp_path / "data" / "draft_saved" / FOLDER_KEY).is_dir()
    assert not (tmp_path / "automation").exists()

    book = openpyxl.load_workbook(tmp_path / "price_inventory.xlsx", read_only=True)
    try:
        rows = list(book.active.iter_rows(values_only=True))
    finally:
        book.close()
    assert rows == [("型号", "价格", "库存"), (MODEL, None, None)]


def test_price_and_stock_are_requested_before_source_files(tmp_path: Path) -> None:
    _write_workbook(tmp_path / "price_inventory.xlsx", None, None)
    _add_source_files(tmp_path, pdfs=0, images=0)

    result = onboard_product(tmp_path, MODEL)

    assert result.status == "NEEDS_PRICE_STOCK"
    assert result.checks["price_present"] is False
    assert result.checks["stock_present"] is False
    assert "价格和库存" in result.next_action


def test_one_missing_inventory_value_is_the_only_next_action(tmp_path: Path) -> None:
    _write_workbook(tmp_path / "price_inventory.xlsx", 120, None)
    _add_source_files(tmp_path, pdfs=0, images=0)

    result = onboard_product(tmp_path, MODEL)

    assert result.status == "NEEDS_PRICE_STOCK"
    assert "库存" in result.next_action
    assert "价格和库存" not in result.next_action
    assert "PDF" not in result.next_action


def test_missing_pdf_is_the_only_next_source_action(tmp_path: Path) -> None:
    _write_workbook(tmp_path / "price_inventory.xlsx", 120, 8)
    _add_source_files(tmp_path, pdfs=0, images=4)

    result = onboard_product(tmp_path, MODEL)

    assert result.status == "NEEDS_SOURCE_FILES"
    assert "PDF" in result.next_action
    assert "照片" not in result.next_action


def test_missing_photos_reports_exact_remaining_count(tmp_path: Path) -> None:
    _write_workbook(tmp_path / "price_inventory.xlsx", 120, 8)
    _add_source_files(tmp_path, pdfs=1, images=2)

    result = onboard_product(tmp_path, MODEL)

    assert result.status == "NEEDS_SOURCE_FILES"
    assert "2张" in result.next_action


def test_complete_inputs_are_ready_to_upload(tmp_path: Path) -> None:
    _write_workbook(tmp_path / "price_inventory.xlsx", 120, 8)
    source = _add_source_files(tmp_path)

    result = onboard_product(tmp_path, MODEL)

    assert result.ok is True
    assert result.status == "READY_TO_UPLOAD"
    assert result.paths["source_directory"] == str(source.resolve())


def test_repeat_run_preserves_business_files_and_does_not_duplicate_row(
    tmp_path: Path,
) -> None:
    _write_workbook(tmp_path / "price_inventory.xlsx", 120, 8)
    source = _add_source_files(tmp_path)
    protected = source / "用户原图.jpg"
    protected.write_bytes(b"do-not-change")

    first = onboard_product(tmp_path, MODEL)
    second = onboard_product(tmp_path, MODEL)

    assert first.status == second.status == "READY_TO_UPLOAD"
    assert protected.read_bytes() == b"do-not-change"
    book = openpyxl.load_workbook(tmp_path / "price_inventory.xlsx", read_only=True)
    try:
        models = [row[0] for row in book.active.iter_rows(min_row=2, values_only=True)]
    finally:
        book.close()
    assert models == [MODEL]


def test_new_model_does_not_modify_other_product_directories(tmp_path: Path) -> None:
    other = tmp_path / "data" / "draft_saved" / "OTHER-MODEL"
    other.mkdir(parents=True)
    protected = other / "keep.pdf"
    protected.write_bytes(b"protected-business-data")

    result = onboard_product(tmp_path, MODEL)

    assert result.status == "NEEDS_PRICE_STOCK"
    assert protected.read_bytes() == b"protected-business-data"
    assert other.is_dir()
